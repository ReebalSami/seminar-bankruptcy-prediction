#!/usr/bin/env python3
"""
Create comprehensive feature mapping Excel file.

Shows:
1. All original feature names for each dataset
2. Semantic mappings (where they exist)
3. Feature descriptions
4. VIF status

Author: Generated
Date: November 13, 2025
"""

import sys
from pathlib import Path
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

import pandas as pd
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

print("=" * 80)
print("CREATING COMPLETE FEATURE MAPPING EXCEL")
print("=" * 80)

# Load datasets to get actual feature names
print("\n[1/5] Loading datasets...")
df_polish = pd.read_parquet(PROJECT_ROOT / 'data' / 'processed' / 'poland_clean_full.parquet')
df_american = pd.read_parquet(PROJECT_ROOT / 'data' / 'processed' / 'american' / 'american_clean.parquet')
df_taiwan = pd.read_parquet(PROJECT_ROOT / 'data' / 'processed' / 'taiwan' / 'taiwan_clean.parquet')

print(f"  Polish: {df_polish.shape}")
print(f"  American: {df_american.shape}")
print(f"  Taiwan: {df_taiwan.shape}")

# Get feature columns
polish_features = [c for c in df_polish.columns if c.startswith('Attr')]
american_features = [c for c in df_american.columns if c.startswith('X')]
taiwan_features = [c for c in df_taiwan.columns if c.startswith('F')]

print(f"\n  Polish features: {len(polish_features)}")
print(f"  American features: {len(american_features)}")
print(f"  Taiwan features: {len(taiwan_features)}")

# Load semantic mappings (if exist)
print("\n[2/5] Loading semantic mappings...")
try:
    with open(PROJECT_ROOT / 'results' / '00_feature_mapping' / 'feature_semantic_mapping_FIXED.json') as f:
        semantic_mapping = json.load(f)['semantic_mappings']
    print(f"  ✓ Loaded semantic mappings: {len(semantic_mapping)} concepts")
except:
    print("  ⚠️ No semantic mappings found")
    semantic_mapping = {}

# Load Taiwan metadata
print("\n[3/5] Loading Taiwan metadata...")
try:
    with open(PROJECT_ROOT / 'data' / 'processed' / 'taiwan' / 'taiwan_features_metadata.json') as f:
        taiwan_metadata = json.load(f)
    print(f"  ✓ Loaded Taiwan metadata: {len(taiwan_metadata)} features")
except:
    print("  ⚠️ No Taiwan metadata found")
    taiwan_metadata = {}

# Load VIF results
print("\n[4/5] Loading VIF results...")
vif_results = {}

try:
    vif_polish = pd.read_csv(PROJECT_ROOT / 'results' / 'script_outputs' / '01_polish' / '10d_remediation_save' / 'vif_all_features.csv')
    vif_results['polish'] = dict(zip(vif_polish['feature'], vif_polish['vif']))
    print(f"  ✓ Polish VIF: {len(vif_results['polish'])} features")
except Exception as e:
    print(f"  ⚠️ No Polish VIF found: {e}")
    vif_results['polish'] = {}

try:
    vif_american = pd.read_csv(PROJECT_ROOT / 'results' / 'script_outputs' / '02_american' / '02b_vif_remediation' / 'vif_all_features.csv')
    vif_results['american'] = dict(zip(vif_american['feature'], vif_american['vif']))
    print(f"  ✓ American VIF: {len(vif_results['american'])} features")
except:
    print("  ⚠️ No American VIF found")
    vif_results['american'] = {}

try:
    vif_taiwan = pd.read_csv(PROJECT_ROOT / 'results' / 'script_outputs' / '03_taiwan' / '02b_vif_remediation' / 'vif_all_features.csv')
    vif_results['taiwan'] = dict(zip(vif_taiwan['feature'], vif_taiwan['vif']))
    print(f"  ✓ Taiwan VIF: {len(vif_results['taiwan'])} features")
except:
    print("  ⚠️ No Taiwan VIF found")
    vif_results['taiwan'] = {}

# Create mappings
print("\n[5/5] Creating Excel file...")

# Create reverse semantic mapping (feature -> concept)
feature_to_semantic = {}
for concept, mapping in semantic_mapping.items():
    for dataset, features in mapping.items():
        for feat in features:
            if dataset not in feature_to_semantic:
                feature_to_semantic[dataset] = {}
            feature_to_semantic[dataset][feat] = concept

# Build dataframes
def create_feature_df(features, dataset_name, vif_dict, metadata=None):
    """Create feature dataframe for a dataset."""
    data = []
    for feat in sorted(features):
        row = {
            'Original_Feature_Name': feat,
            'Semantic_Mapping': feature_to_semantic.get(dataset_name.lower(), {}).get(feat, ''),
            'VIF': vif_dict.get(feat, ''),
            'VIF_Status': ''
        }
        
        # Add VIF status
        if row['VIF'] != '':
            vif_val = row['VIF']
            if vif_val < 5:
                row['VIF_Status'] = 'LOW (<5)'
            elif vif_val < 10:
                row['VIF_Status'] = 'MODERATE (5-10)'
            else:
                row['VIF_Status'] = 'HIGH (≥10)'
        
        # Add description for Taiwan
        if metadata and feat in metadata:
            # Handle both string and dict metadata
            if isinstance(metadata[feat], dict):
                row['Description'] = metadata[feat].get('original_name', '')
            else:
                row['Description'] = metadata[feat]
        
        data.append(row)
    
    return pd.DataFrame(data)

# Create workbook
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# Polish sheet
ws_polish = wb.create_sheet("Polish Features")
df_polish_features = create_feature_df(polish_features, 'polish', vif_results['polish'])
for r in dataframe_to_rows(df_polish_features, index=False, header=True):
    ws_polish.append(r)

# American sheet
ws_american = wb.create_sheet("American Features")
df_american_features = create_feature_df(american_features, 'american', vif_results['american'])
for r in dataframe_to_rows(df_american_features, index=False, header=True):
    ws_american.append(r)

# Taiwan sheet
ws_taiwan = wb.create_sheet("Taiwan Features")
df_taiwan_features = create_feature_df(taiwan_features, 'taiwan', vif_results['taiwan'], taiwan_metadata)
for r in dataframe_to_rows(df_taiwan_features, index=False, header=True):
    ws_taiwan.append(r)

# Summary sheet
ws_summary = wb.create_sheet("Summary", 0)
summary_data = [
    ['Dataset Feature Mapping Summary', '', '', ''],
    ['', '', '', ''],
    ['Dataset', 'Total Features', 'Semantically Mapped', 'VIF < 5'],
    ['Polish', len(polish_features), 
     sum(1 for f in polish_features if feature_to_semantic.get('polish', {}).get(f)),
     sum(1 for f in polish_features if vif_results['polish'].get(f, 999) < 5)],
    ['American', len(american_features),
     sum(1 for f in american_features if feature_to_semantic.get('american', {}).get(f)),
     sum(1 for f in american_features if vif_results['american'].get(f, 999) < 5)],
    ['Taiwan', len(taiwan_features),
     sum(1 for f in taiwan_features if feature_to_semantic.get('taiwan', {}).get(f)),
     sum(1 for f in taiwan_features if vif_results['taiwan'].get(f, 999) < 5)],
    ['', '', '', ''],
    ['Semantic Concepts', '', '', ''],
    ['Concept', 'Polish Features', 'American Features', 'Taiwan Features'],
]

for concept, mapping in sorted(semantic_mapping.items()):
    summary_data.append([
        concept,
        ', '.join(mapping.get('polish', [])),
        ', '.join(mapping.get('american', [])),
        ', '.join(mapping.get('taiwan', []))
    ])

for row in summary_data:
    ws_summary.append(row)

# Style headers
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

for ws in [ws_summary, ws_polish, ws_american, ws_taiwan]:
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Auto-width
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

# Save
output_file = PROJECT_ROOT / 'COMPLETE_FEATURE_MAPPING.xlsx'
wb.save(output_file)

print(f"\n✅ Created: {output_file}")
print(f"\n  Sheets:")
print(f"    1. Summary - Overview of all datasets")
print(f"    2. Polish Features - {len(polish_features)} features")
print(f"    3. American Features - {len(american_features)} features")
print(f"    4. Taiwan Features - {len(taiwan_features)} features")
print("\n  Columns:")
print("    • Original_Feature_Name - Actual feature name in dataset")
print("    • Semantic_Mapping - Mapped concept (if exists)")
print("    • VIF - Variance Inflation Factor")
print("    • VIF_Status - LOW/MODERATE/HIGH")
print("    • Description - (Taiwan only) Original feature description")

print("\n" + "=" * 80)
